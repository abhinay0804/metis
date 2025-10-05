"""
Excel content extractor using openpyxl.
"""

import openpyxl
from typing import Dict, Any, List
import json


class ExcelExtractor:
    """Extract data and metadata from Excel files."""
    
    def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Extract content from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing extracted content
        """
        content = {
            "worksheets": [],
            "metadata": {},
            "sheet_count": 0
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            content["sheet_count"] = len(workbook.worksheets)
            content["metadata"] = {
                "title": workbook.properties.title,
                "author": workbook.properties.creator,
                "subject": workbook.properties.subject,
                "created": str(workbook.properties.created) if workbook.properties.created else None,
                "modified": str(workbook.properties.modified) if workbook.properties.modified else None
            }
            
            for sheet in workbook.worksheets:
                sheet_content = {
                    "sheet_name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "data": [],
                    "formulas": []
                }
                
                # Extract data from all cells
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        sheet_content["data"].append(list(row))
                
                # Extract formulas (if needed)
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f' and cell.value:  # Formula cell
                            sheet_content["formulas"].append({
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })
                
                content["worksheets"].append(sheet_content)
        
        except Exception as e:
            raise Exception(f"Error extracting Excel content: {str(e)}")
        
        return content
